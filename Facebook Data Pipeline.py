import openpyxl
import os
from dotenv import load_dotenv
from openpyxl.styles import Font
from facebook_business.api import FacebookAdsApi
from facebook_business.adobjects.adaccount import AdAccount

# Load environement variables
load_dotenv()

# Getting Acces ID from env_vars
my_app_id = os.environ.get("my_app_id")
my_app_secret = os.environ.get("my_app_secret")
my_access_token = os.environ.get("my_access_token")

# Initialitiong connection with Facebook
FacebookAdsApi.init(my_app_id, my_app_secret, my_access_token)

# Connecting to accounts
laSalle_account = AdAccount(os.environ.get("laSalle_AdAccount"))
interDec_account = AdAccount(os.environ.get("interDec_AdAccount"))

# Seting-Up my requests
params = {
    'date_preset': "this_quarter",
    "level": "campaign",
    'time_increment': "1",
    "export_format": "csv"
}

fields = ["campaign_name", "reach", "clicks", "spend"]

# Requesting insights
laSalle_insights = laSalle_account.get_insights(params=params, fields=fields)
interDec_insights = interDec_account.get_insights(params=params, fields=fields)

# Transform insights into a list of dict
laSalle_insights = [dict(x) for x in laSalle_insights]
interDec_insights = [dict(x) for x in interDec_insights]

# Print result
print(laSalle_insights)
print(interDec_insights)

# Write insights into file

# Connect to file
workbook = os.environ.get("my_workbook")
path = os.getcwd()
wb = openpyxl.load_workbook(f"{path}\\{workbook}")

# Refresh data
old = wb["FacebookLastWeek"]
wb.remove(old)
ws = wb.create_sheet("FacebookLastWeek", 0)

dest_filename = f"{path}\\{workbook}"

# Write headers
headers = ("Campaign Name", "Clicks", "Reach", "Spent", "Date",)
ws.append(headers)

# Formating header
ws.column_dimensions["A"].width = 44
bold_font = Font(bold=True, italic=True, size=12)
for cell in ws["1:1"]:
    cell.font = bold_font


# Write new data
for i in range(0, len(laSalle_insights)):
    ws.append((
        laSalle_insights[i]["campaign_name"],
        laSalle_insights[i]["clicks"],
        laSalle_insights[i]["reach"],
        laSalle_insights[i]["spend"],
        laSalle_insights[i]['date_start'],
        ))

for i in range(0, len(interDec_insights)):
    ws.append((
        interDec_insights[i]["campaign_name"],
        interDec_insights[i]["clicks"],
        interDec_insights[i]["reach"],
        interDec_insights[i]["spend"],
        interDec_insights[i]['date_start'],
        ))

# Formating spent column
for i in range(2, 20):
    _cell = ws.cell(row=i, column=4)
    _cell.number_format = "$#,##0.00"

# Saving
wb.save(dest_filename)
